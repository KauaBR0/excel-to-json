import openpyxl
import json
import re
from datetime import datetime, time


def converter_valor(valor):
  if isinstance(valor, (datetime, time)):
    return valor.isoformat()
  elif isinstance(valor, str):
    return re.sub(r'_x000D_\n?', '', valor)
  return valor


caminho_planilha = 'planilha.xlsx'
workbook = openpyxl.load_workbook(caminho_planilha)
worksheet = workbook.active

cabecalhos = []
for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)):
  if cell:
    cabecalhos.append(cell)

dados = []
for row in worksheet.iter_rows(min_row=2, values_only=True):
  dados.append({cabecalhos[i]: converter_valor(v) for i, v in enumerate(row)})

json_data = json.dumps(dados, indent=4, ensure_ascii=False)

with open('dados.json', 'w', encoding='utf-8') as arquivo:
  arquivo.write(json_data)
